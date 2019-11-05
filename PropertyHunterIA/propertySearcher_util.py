# download_image
import requests
import os

#  classify_image
from clustering_vae import * # 10.18 original
from pathlib import Path
import pandas as pd

# email
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.application import MIMEApplication
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email.mime.image import MIMEImage
from email import encoders
# from propertyguru_v1 import num_result, property_title
from email.header import Header
import time

# text_classifier
import pandas as pd
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.cluster import KMeans

# excel editor
from openpyxl import load_workbook
from openpyxl import worksheet

import tagui as t
import sys

SENDER = 'EMAIL'
PASSWORD = 'PASSWORD'

def wait_for_pageload(selector):
    wait_status = 0
    for loop_wait in range(1, 20):
        print(f"{loop_wait}. waiting for page to appear. wait for 1s...")
        if t.present(selector):
            wait_status = 1
            break
        else:
            t.wait(1)
    print("bp801.2 wait_status = {}".format(wait_status))

def wait_for_mainpageload(selector):
    wait_status = 0
    result = 1
    for loop_wait in range(1, 15):
        print(f"{loop_wait}. waiting for page to appear. wait for 1s...")
        if t.present(selector):
            wait_status = 1
            break
        else:
            t.wait(1)
            if loop_wait == 14:
                result = 0
    return result

    print("bp801.2 wait_status = {}".format(wait_status))


def hover_and_read(selector):
    t.hover(selector)
    str = t.read(selector)
    return str


def read_if_present(selector):
    str = ""
    if t.present(selector):
        str = t.read(selector)
    return str


def download_image(id, image1, image2, image3):
    """

    :param id: property id
    :param image1: image1 url
    :param image2: image2 url
    :param image3: image3 url
    :return:
    """
    print('======== downloading image start==========')

    os.makedirs('./image/', exist_ok=True)
    def request_download(id, IMAGE_URL,image):
        r = requests.get(IMAGE_URL)
        with open(f'./image/{id} {image}.png', 'wb') as f:
            f.write(r.content)

    for (n,i,m) in zip(id, image1, range(1, len(id) + 1)):
        try:
            request_download(n, i, "view1")
            print(f'{m} {n} view1 saved')
        except:
            print(f"{m} {n}Please enter a valid URL")

    for (n,i,m) in zip(id, image2, range(1, len(id) + 1)):
        try:
            request_download(n, i, "view2")
            print(f'{m} {n} view2 saved')
        except:
            print(f"{m} {n} Please enter a valid URL")

    for (n,i,m) in zip(id, image3, range(1, len(id) + 1)):
        try:
            request_download(n, i, "view3")
            print(f'{m} {n} view3 saved')
        except:
            print(f"{m} {n} Please enter a valid URL")

    print('======== downloading image end ==========')

# read excel will hit error when read id and nan, differ from read from dataframe
def classify_image(data, prefer1, prefer2, prefer3):
    """
    :param data: property dataframe
    :param prefer1:
    :param prefer2:
    :param prefer3:
    :return: filtered_name, filtered_cluster
    """
    id = data.id.values.tolist()
    imagepath = Path('./image/')

    # fit model
    shortlist = dict()
    shortlist_cluster = dict()
    for i in id:
        if i != '':
            print(f'======== start clustering {i} ==========')
            # get cluster
            try:
                view1 = predict_image_cluster(imagepath / f'{i} view1.png')
            except:
                view1 = 'N'
                print('no image1 found')
            try:
                view2 = predict_image_cluster(imagepath / f'{i} view2.png')
            except:
                view2 = 'N'
                print('no image2 found')
            try:
              view3 = predict_image_cluster(imagepath / f'{i} view3.png')
            except:
                view3= 'N'
                print('no image3 found')

            # add cluster in dict
            shortlist_cluster[i] = f'{view1}|{view2}|{view3}'

            # get score
            score_list = [view1,view2,view3]
            print(score_list)
            score1 = score_list.count(prefer1)
            score2 = score_list.count(prefer2)
            score3 = score_list.count(prefer3)
            print(f'{i} preference 1 score is {score1}')
            print(f'{i} preference 2 score is {score2}')
            print(f'{i} preference 3 score is {score3}')

            # add score in dict
            shortlist[i] = score1 + score2 + score3
            print(f'{i} total preference score is  {shortlist[i]}')

        else:
            print('nan value')

    print('======== sorting ==========')
    # sort by largest score
    a = sorted(shortlist.items(), key=lambda item:item[1], reverse=True)
    print(f'sort by score {a}')

    # get top 5 score
    b = a[0:5]
    print(f'top 5 score {b}')

    # get top 5 property id list
    filtered_id = list()
    for i in b:
        filtered_id.append(i[0])
    print(f'top 5 filter_id  {filtered_id}')

    # get top 5 cluster list
    filtered_cluster = list()
    for key in filtered_id:
       filtered_cluster.append(shortlist_cluster[key])
    print(f'top 5 filtered_cluster {filtered_cluster}')

    return filtered_id , filtered_cluster


def download_pdf(property_title, pdf_link,pdf_id):
    print('======== downloading pdf start ==========')
    filename = list()
    for n, i, x in zip(pdf_link,property_title,pdf_id):
        if n != '' and i != '':
            pdf = f'https://www.propertyguru.com.sg/pdf/{n[-8:]}.pdf'
            print(f'downloading {i} {pdf}')
            r = requests.get(pdf, allow_redirects=True)
            # filename = get_filename_from_cd(r.headers.get('content-disposition'))
            open(f'{i} ID[{x}].pdf', 'wb').write(r.content)
            filename.append(f'{i} ID[{x}].pdf')
    print(filename)
    print('======== downloading pdf end ==========')

    return filename


def mail_shortlist(input_email, input_name, pdf_filename):
    print('============ compiling shortlist email ============')

    localtime = time.asctime(time.localtime(time.time()))

    sender = SENDER
    password = PASSWORD
    receiver = input_email

    smtpserver = smtplib.SMTP('smtp.gmail.com',587)
    smtpserver.ehlo()
    smtpserver.starttls()
    smtpserver.ehlo()
    smtpserver.login(sender,password)

    msg = MIMEMultipart('mixed')
    msg['Subject'] = 'Shortlisted Properties'
    msg['From'] = 'isspropertyagent@gmail.com <isspropertyagent@gmail.com>'
    msg['To'] = input_email

    text = f"Dear {input_name}," \
        f"\n" \
        f"\nWe have found some properties which you may be interested in!" \
        f"\n" \
        f"\nPlease view the attachments and let the Property Hunter Chat-Bot know which property you like to view by specifying the 8 digit property ID (e.g. I like to view property 12345678)." \
        f"\n" \
        f"\nCheers," \
        f"\nProperty Hunter" \
        f"\n" \
        f"\n" \
        f"\nThis is an automatically generated email. Please do not reply. {localtime}"

    text_plain = MIMEText(text,'plain', 'utf-8')
    msg.attach(text_plain)

    # sendimagefile=open(r'E:\PycharmProjects\untitled1\test.png','rb').read()
    # image = MIMEImage(sendimagefile)
    # image.add_header('Content-ID','<image1>')
    # image["Content-Disposition"] = 'attachment; filename="propertyguru_image.png"'
    # msg.attach(image)

    with open('Property Detail.xlsx', "rb") as attachment:
        part = MIMEBase("application", "octet-stream")
        part.set_payload(attachment.read())
    # Encode file in ASCII characters to send by email
    encoders.encode_base64(part)
    # Add header as key/value pair to attachment part
    part.add_header(
        "Content-Disposition",
        f"attachment; filename= {'Property Detail.xlsx'}",
    )
    msg.attach(part)

    for n in pdf_filename:
        with open(n, "rb") as attachment:
            part = MIMEBase("application", "octet-stream")
            part.set_payload(attachment.read())
        # Encode file in ASCII characters to send by email
        encoders.encode_base64(part)
        # Add header as key/value pair to attachment part
        part.add_header(
            "Content-Disposition",
            f"attachment; filename= {n}",
        )
        msg.attach(part)

    smtpserver.sendmail(sender,receiver,msg.as_string())
    smtpserver.close()
    print('============ shortlist email sent ============')


def mail_subscription(input_email, input_name, pdf_filename):
    print('============ compiling subscription email ============')

    localtime = time.asctime(time.localtime(time.time()))

    sender = SENDER
    password = PASSWORD
    receiver = input_email

    smtpserver = smtplib.SMTP('smtp.gmail.com',587)
    smtpserver.ehlo()
    smtpserver.starttls()
    smtpserver.ehlo()
    smtpserver.login(sender,password)

    msg = MIMEMultipart('mixed')
    msg['Subject'] = 'New listings'
    msg['From'] = 'isspropertyagent@gmail.com <isspropertyagent@gmail.com>'
    msg['To'] = input_email

    text = f"Dear {input_name}," \
        f"\n" \
        f"\nWe have found some new listings which you may be interested in!" \
        f"\n" \
        f"\n" \
        f"\nPlease view the attachments and let the Property Hunter Chat-Bot know which property you like to view by specifying the 8 digit property ID (e.g. I like to view property 12345678)." \
        f"\n" \
        f"\nCheers," \
        f"\nProperty Hunter" \
        f"\n" \
        f"\n" \
        f"\nThis is an automatically generated email. Please do not reply. {localtime}"

    text_plain = MIMEText(text,'plain', 'utf-8')
    msg.attach(text_plain)

    with open('Property Monitor.xlsx', "rb") as attachment:
        part = MIMEBase("application", "octet-stream")
        part.set_payload(attachment.read())
    # Encode file in ASCII characters to send by email
    encoders.encode_base64(part)
    # Add header as key/value pair to attachment part
    part.add_header(
        "Content-Disposition",
        f"attachment; filename= {'Property Monitor.xlsx'}",
    )
    msg.attach(part)

    for n in pdf_filename:
        with open(n, "rb") as attachment:
            part = MIMEBase("application", "octet-stream")
            part.set_payload(attachment.read())
        # Encode file in ASCII characters to send by email
        encoders.encode_base64(part)
        # Add header as key/value pair to attachment part
        part.add_header(
            "Content-Disposition",
            f"attachment; filename= {n}",
        )
        msg.attach(part)

    smtpserver.sendmail(sender,receiver,msg.as_string())
    smtpserver.close()
    print('============ subscription email sent ============')


def mail_confirmation(input_email, input_name, view_project, view_date, view_addr, view_time, view_price):
    """

    :param input_email:
    :param input_name:
    :param view_project:
    :param view_date:
    :param view_addr:
    :param view_time:
    :param view_price:
    :return:
    """

    print('============ compiling confirmation email ============')
    localtime = time.asctime(time.localtime(time.time()))

    sender = SENDER
    password = PASSWORD
    receiver = input_email

    smtpserver = smtplib.SMTP('smtp.gmail.com',587)
    smtpserver.ehlo()
    smtpserver.starttls()
    smtpserver.ehlo()
    smtpserver.login(sender,password)

    msg = MIMEMultipart('mixed')
    msg['Subject'] = 'Viewing Appointment'
    msg['From'] = 'isspropertyagent@gmail.com <isspropertyagent@gmail.com>'
    msg['To'] = input_email

    text = f"Dear {input_name}," \
        f"\n" \
        f"\nPlease note that your viewing appointment for the unit at {view_project} on {view_date} is confirmed. Please see the details of the appointment below." \
        f"\n" \
        f"\nAddress: {view_addr}" \
        f"\nDate: {view_date}" \
        f"\nTime: {view_time}" \
        f"\nFair Estimated Market Price: S$ {view_price}" \
        f"\n" \
        f"\nCheers," \
        f"\nProperty Hunter" \
        f"\n" \
        f"\n" \
        f"\nThis is an automatically generated email. Please do not reply." \
        f"\n{localtime}"
    text_plain = MIMEText(text,'plain', 'utf-8')
    msg.attach(text_plain)

    smtpserver.sendmail(sender,receiver,msg.as_string())
    smtpserver.close()
    print('============ confirmation email sent ============')

def mail_notfound(input_email, input_name,location, size, budget, bedroom,level):

    print('============ compiling no property found email ============')
    localtime = time.asctime(time.localtime(time.time()))

    sender = SENDER
    password = PASSWORD
    receiver = input_email

    smtpserver = smtplib.SMTP('smtp.gmail.com',587)
    smtpserver.ehlo()
    smtpserver.starttls()
    smtpserver.ehlo()
    smtpserver.login(sender,password)

    msg = MIMEMultipart('mixed')
    msg['Subject'] = 'No Matching Properies Found'
    msg['From'] = 'isspropertyagent@gmail.com <isspropertyagent@gmail.com>'
    msg['To'] = input_email

    text = f"Dear {input_name}," \
        f"\n" \
        f"\nAt the moment we could not find any property that matches your requirements below. You may want to try with other requirements. " \
        f"\n" \
        f"\nLocation: {location}" \
        f"\nSize: {size} sqft" \
        f"\nBudget: ${budget}" \
        f"\nNumber of bedroom: {bedroom[0].replace('.0','')}"\
        f"\nFloor level: {level[0]}" \
        f"\n" \
        f"\nCheers," \
        f"\nProperty Hunter" \
        f"\n" \
        f"\n" \
        f"\nThis is an automatically generated email. Please do not reply. {localtime}"

    text_plain = MIMEText(text,'plain', 'utf-8')
    msg.attach(text_plain)

    smtpserver.sendmail(sender,receiver,msg.as_string())
    smtpserver.close()
    print('============ no property found email sent ============')



def classify_text(data, num_clusters, features_per):
    """

    :param data:
    :param num_clusters: select num of clusters
    :param features_per: select num of features per cluster to display
    :return: features_selected
    """

    print('============ text classification start ============')
    # data = pd.read_excel(filename)
    x = data['id'].tolist()
    y = data['feature'].tolist()
    print(x)
    print(y)

    """# tf-idf Vectorizer"""
    tfidf_vectorizer = TfidfVectorizer(max_features=200000,
                                       stop_words='english',
                                       use_idf=True,
                                      max_df=0.8,min_df=0.2,ngram_range=(1,1))

    try:
        tfidf_matrix = tfidf_vectorizer.fit_transform(y)#fit the vectorizer to synopses
    except:
        return [''] * len(x)
    terms = tfidf_vectorizer.get_feature_names()

    """# k-means classification"""
    km = KMeans(n_clusters=num_clusters)
    km.fit(tfidf_matrix)
    clusters = km.labels_.tolist()

    """# result interpretation"""
    # build dataframe
    left = pd.DataFrame()
    left['id'] = x
    left['cluster'] = km.labels_
    print(left)

    # Top keywords per row
    print("Top keywords per row:")
    order_centroids = km.cluster_centers_.argsort()[:, ::-1]
    terms = tfidf_vectorizer.get_feature_names()
    centers = list()
    cluster = list()


    for i in range(num_clusters):
        top_ten_words = [terms[ind] for ind in order_centroids[i, :features_per]]
        print("Cluster {}: {}".format(i,' '.join(top_ten_words)))
        centers.append(' '.join(top_ten_words))
        cluster.append(i)

    right = pd.DataFrame({'centers':centers},index=cluster)
    print(right)

    results = left.join(right, on='cluster')
    print(results)

    left1= [str(x) for x in results['cluster'].tolist()]
    right1 = results['centers'].tolist()

    features = [right1[i] for i in range(min(len(left1),len(right1)))]
    print(features)

    features_selected = list()
    for n in features:
        n = n.replace(' features', '')
        n = n.replace(' key', '')
        n = n.replace(' room', '')
        features_selected.append(n)
    print(features_selected)
    print('============ text classification end ============')

    return features_selected


def edit_excel(excelfile):

    wb = load_workbook(excelfile)
    ws = wb[wb.sheetnames[0]]
    ws.freeze_panes = 'B1'
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 9
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 9
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['H'].width = 10
    ws.column_dimensions['I'].width = 23
    ws.column_dimensions['J'].width = 9
    ws.column_dimensions['K'].width = 45
    ws.column_dimensions['L'].width = 10
    ws.column_dimensions['M'].width = 20
    ws.column_dimensions['O'].width = 6
    ws.column_dimensions['P'].width = 100
    wb.save('Property Detail.xlsx')


def getPropertyDetails(id):
    print(f'testing ----------->{id} (typre{type(id)}')
    df = pd.read_excel('Property Detail.xlsx')
    print(df)
    df = df[df['id'].isin([id])]
    print(df)
    i = df.index.values[0]
    print(i)
    projectName = df.loc[[i], ['property_title']].to_string(header=False, index=False)
    projectName = projectName.strip()
    print(f'[{projectName}]',type(projectName))
    district = df.loc[[i], ['region']].to_string(header=False, index=False)
    print(district, type(district))
    floorArea = df.loc[[i], ['area']].to_string(header=False, index=False)
    floorArea = int(floorArea.replace('sqft','').strip())
    print(floorArea, type(floorArea))
    floorHeight = df.loc[[i], ['floor']].to_string(header=False, index=False)
    floorHeight  = floorHeight.strip()
    print(floorHeight, type(floorHeight))
    address = df.loc[[i], ['address']].to_string(header=False, index=False)
    address = address.strip()
    print(address, type(address))

    cluster1 = df.loc[[i], ['image']].to_string(header=False, index=False)[1]
    print(cluster1, type(cluster1))
    cluster2 = df.loc[[i], ['image']].to_string(header=False, index=False)[3]
    print(cluster2, type(cluster2))
    cluster3 = df.loc[[i], ['image']].to_string(header=False, index=False)[5]
    print(cluster3, type(cluster3))

    if district == ' Boat Quay / Raffles Place / Marina':
        D = '1'
    if district == ' Chinatown / Tanjong Pagar':
        D = '2'
    if district == ' Alexandra / Commonwealth':
        D = '3'
    if district ==  'Harbourfront / Telok Blangah':
        D = '4'
    if district == ' Buona Vista / West Coast / Clementi New Town':
        D = '5'
    if district == ' City Hall / Clarke Quay':
        D = '6'
    if district == ' Beach Road / Bugis / Rochor':
        D = '7'
    if district == ' Farrer Park / Serangoon Rd':
        D = '8'
    if district == ' Orchard / River Valley':
        D = '9'
    if district == ' Tanglin / Holland / Bukit Timah':
        D = '10'
    if district == ' Newton / Novena':
        D = '11'
    if district == ' Balestier / Toa Payoh':
        D = '12'
    if district == ' Macpherson / Potong Pasir':
        D = '13'
    if district == ' Eunos / Geylang / Paya Lebar':
        D = '14'
    if district == ' East Coast / Marine Parade':
        D = '15'
    if district == ' Bedok / Upper East Coast':
        D = '16'
    if district == ' Changi Airport / Changi Village':
        D = '17'
    if district == ' Pasir Ris / Tampines':
        D = '18'
    if district == ' Hougang / Punggol / Sengkang':
        D = '19'
    if district == ' Ang Mo Kio / Bishan / Thomson':
        D = '20'
    if district == ' Clementi Park / Upper Bukit Timah':
        D = '21'
    if district == ' Boon Lay / Jurong / Tuas':
        D = '22'
    if district == ' Dairy Farm / Bukit Panjang / Choa Chu Kang':
        D = '23'
    if district == ' Lim Chu Kang / Tengah':
        D = '24'
    if district == ' Admiralty / Woodlands':
        D = '25'
    if district == ' Mandai / Upper Thomson':
        D = '26'
    if district == ' Sembawang / Yishun':
        D = '27'
    if district == ' Seletar / Yio Chu Kang':
        D = '28'
    D = int(D)

    return projectName, D, floorArea, floorHeight, address, cluster1, cluster2, cluster3

if __name__ == "__main__":
    getPropertyDetails('21960546')


